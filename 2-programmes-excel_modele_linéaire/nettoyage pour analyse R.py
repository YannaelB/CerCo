#On importe les bibliothèques nécessaire à l'écriture du programme

import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
import matplotlib.pyplot as plt
import os
from openpyxl.styles import Color,PatternFill
from numpy.linalg import solve
from collections import Counter
import statistics as st
from scipy import stats


 #========================
        #Ouvrir les fichiers excel à traiter 
 #========================


def ouvrirXLSX(filepath) :  #cette fonction permet d'ouvrir un fichier excel selon un chemin donné
    wb = load_workbook(filepath)
    return wb,wb.sheetnames

#Cette fonction permet de parcourir le fichier excel et de le copier/coller dans un tableau pour le manipuler plus simplement. 
def lireFeuille(wb,nom_feuille,prem_ligne = 1,der_ligne=np.inf):   
    ws = wb[nom_feuille]
    tableau = []
    der_ligne = min(der_ligne,ws.max_row)
    for i,row in enumerate(ws.rows,1) : # pour chaque ligne,
        if prem_ligne <= i <= der_ligne :
            ligne = []
            for cell in row : # pour chaque cellule de la ligne,
                ligne.append(cell.value) # lecture de son contenu
            tableau.append(ligne)
    tableau = np.array(tableau) # Conversion en tableau numpy
    shp,typ_don = tableau.shape,tableau.dtype
    return tableau



classeur = "Sujet_N_MutilAnalysis-nettoyé3.1.xlsx" #chemin du fichier créé par "nettoyage excel_v3". L'on traitera désormais le fichier excel allégé car cela va simplifier les algorithmes

# creation d'une liste de tous les tableaux contenants les différents sujets/essais
listeFeuilles = [lireFeuille(ouvrirXLSX(classeur)[0],i) for i in (ouvrirXLSX(classeur)[1])]



 #========================
        #Création d'un tableau en partant de zéro et en prélevant les valeurs utiles dans l'excel
 #========================


#fonction qui retourne la liste des erreurs pour un sujet s, une salle, R et un son n
def etude_stimuli(L,s,R,n):
    nb_hand = []
    nb_head = []
    TR = [] #on crée des listes vides que l'on va remplir petit à petit
    Errh = []
    Errv = []
    Errp = []
    Err3D = []
    ErrHead = []
    for k in range(1,136): #Pour chaque essai du sujet "s"
        if int(L[s][k,2]) == R: #s'ils sont dans la room R
            if int(L[s][k,3]) == n: #s'ils ont pour stimsound le numéro n
                if np.isnan(float(L[s][k,8])) == False: #on vérifie que les valeurs existent
                    TR.append(float(float(L[s][k,7]))) #ajoute dans une liste le TR du stimuli en question 
                    nb_hand.append(float(float(L[s][k,6])))
                    nb_head.append(float(float(L[s][k,5])))
                    Errh.append(abs(float(L[s][k,9])))
                    Errv.append(abs(float(L[s][k,10])))
                    Errp.append(abs(float(L[s][k,11])))
                    Err3D.append(abs(float(L[s][k,8])))
                if np.isnan(float(L[s][k,12])) == False:
                    ErrHead.append(abs(float(L[s][k,12])))

    return nb_head,nb_hand,TR,Err3D,Errh,Errv,Errp,ErrHead #la fonction renvoie des listes

#fonction qui retourne la moyenne des erreurs pour un sujet s, une salle R, des sons numéro n1 à n2
def in_cong(L,s,R,n1,n2):
    nb_hand1 = []
    nb_head1 = [] #comme jusqu'à maintenant, on crée des listes vides que l'on va remplir
    TR1 = []
    Errh1 = []
    Errv1 = [] 
    Errp1 = []
    Err3D1 = []
    ErrHead1 = []
    for i in range(n1,n2+1): #on va appeler la fonction précédente sur l'interval de stimsound [n1,n2+1[ => en jouant avec R, cela correspondra aux sons congruents/incongruents
        nb_head,nb_hand,TR,Err3D,Errh,Errv,Errp,ErrHead = etude_stimuli(L,s,R,i) #c'est la fonction défini juste au dessus
        for j in range(len(Errh)): #la moyenne de listes de listes n'a pas de sens donc l'on vient ajouter chaque valeur des petites listes erreurs dans des plus grand listes contenant les erreurs de n1 à n2  // on aurait peut-être aussi pu se contenter de modifier le format d'une liste de liste en une simple longue liste mais les sous listes ne font pas toutes la mêmes tailles donc un np.reshape ne convient pas
            nb_head1.append(nb_head[j])
            nb_hand1.append(nb_hand[j])
            TR1.append(TR[j])
            Errh1.append(Errh[j])
            Errv1.append(Errv[j])
            Errp1.append(Errp[j])
            Err3D1.append(Err3D[j])
        for k in range(len(ErrHead)):
            ErrHead1.append(ErrHead[k])
    return st.mean(nb_head1),st.mean(nb_hand1),st.mean(TR1),st.mean(Err3D1),st.mean(Errh1),st.mean(Errv1),st.mean(Errp1),st.mean(ErrHead1) #on retourne les moyennes des grandes listes
            

    
#fonction qui crée un tableau pour le modèle linéaire de R
def création_tab_R(L):
    n=len(L) #nombre de sujets
    T = np.zeros([9*n+1,11])  #créé un tableau de 0, on veut 11 colonnes et 9lignes par sujet car 3 scènes avec 3 types de sons. La ligne en plus permet de rajouter les titres
    w = 1
    for i in range(n):#pour chaque sujet
        T[w:w+9,0] = i+1
        for k in range(1,4): #Pour les 3 salles
            for l in range(1,4):  #parcours les sons de la plage/eglise/neutres
                if k == 1:
                    if l == 1:
                        a,b,c,d,e,f,g,h=in_cong(L,i,8,1,9) #retourne la moyenne des erreurs de sons de la plage dans la plage
                        T[w,1:11] = k,l,a,b,c,d,e,f,g,h  #insert la moyenne des erreurs des sons de la plage dans plage
                    elif l == 2:
                        a,b,c,d,e,f,g,h=in_cong(L,i,8,10,18) #moyenne des sons de l'église dans la plage
                        T[w,1:11] = k,l,a,b,c,d,e,f,g,h
                    elif l == 3:
                        a,b,c,d,e,f,g,h=in_cong(L,i,8,19,21) #c'est la fonction défini juste au dessus
                        T[w,1:11] = k,l,a,b,c,d,e,f,g,h
                elif k == 2:
                    if l == 1:
                        a,b,c,d,e,f,g,h=in_cong(L,i,10,1,9) #moyenne des sons de la plage dans l'église
                        T[w,1:11] = k,l,a,b,c,d,e,f,g,h
                    elif l == 2:
                        a,b,c,d,e,f,g,h =in_cong(L,i,10,10,18)
                        T[w,1:11] = k,l,a,b,c,d,e,f,g,h
                    elif l == 3:
                        a,b,c,d,e,f,g,h =in_cong(L,i,10,19,21)
                        T[w,1:11] = k,l,a,b,c,d,e,f,g,h
                elif k == 3:
                    if l == 1:
                        a,b,c,d,e,f,g,h =in_cong(L,i,0,1,9) 
                        T[w,1:11] = k,l,a,b,c,d,e,f,g,h
                    elif l == 2:
                        a,b,c,d,e,f,g,h=in_cong(L,i,0,10,18)
                        T[w,1:11] = k,l,a,b,c,d,e,f,g,h
                    elif l == 3:
                        a,b,c,d,e,f,g,h =in_cong(L,i,0,19,21)
                        T[w,1:11] = k,l,a,b,c,d,e,f,g,h
                w = w+1
                #on a bien 3 if avec 3 sous if qui permettent de représenter les 9 conditions voulu dans le modèle linéaire R
    return T


#Cette fonction permet de transformer la liste de tableau en un fichier excel avec des feuilles
def creationwb(L):
    wb=Workbook()

    ws1 = wb.create_sheet("modèle linéaire") #on crée une feuille dans un fichier excel
        
    # on rempli le tableau cellule par cellule en copiant/collant le tableau crée avec la fonction "création_tab_R"
    for ligne in range (1,L.shape[0]+1) :
        for colonne in range (1,L.shape[1]+1):
            ws1.cell(row= ligne, column = colonne ,value = L[ligne-1][colonne-1] )
            
    ws1.cell(1,1 ,value = "Subject" ) #on vient rajouter les titres des colonnes. On le fait après car le tableau numpy contenait des number float et n'acceptait donc pas les string (chaîne de caractère)
    ws1.cell(1,2 ,value = "Room" )
    ws1.cell(1,3 ,value = "Cong" )
    ws1.cell(1,4 ,value = "nb_head_move" )
    ws1.cell(1,5 ,value = "HAND_numberOfMovement" )
    ws1.cell(1,6 ,value = "HEAD_RT" )
    ws1.cell(1,7 ,value = "Distance_meanSound_3D" )
    ws1.cell(1,8 ,value = "HandSignedError_h" )
    ws1.cell(1,9 ,value = "HandSignedError_v" )
    ws1.cell(1,10 ,value = "HandDepth_signed" )
    ws1.cell(1,11 ,value = "Head_error_orientation" )      
   
    del wb["Sheet"]
    wb.save("Excel_pour_analyze_Rtest2.xlsx")
    return print("l'excel à été créé ;) ")

                
 #========================
        #Appeler les fonctions dernières avec les noms des fichiers à traiter 
 #========================
               
T = création_tab_R(listeFeuilles) #ici l'ordre est très important
creationwb(T)
        

        
        






        
