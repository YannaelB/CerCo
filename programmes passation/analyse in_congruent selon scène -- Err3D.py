#On importe les bibliothèques nécessaire à l'écriture du programme

import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
import matplotlib.pyplot as plt
import os
from openpyxl.styles import Color,PatternFill
from numpy.linalg import solve
from collections import Counter
import statistics as st
from scipy import stats


 #========================
        #Ouvrir les fichiers excel à traiter 
 #========================

def ouvrirXLSX(filepath) :
    wb = load_workbook(filepath)
    return wb,wb.sheetnames

def lireFeuille(wb,nom_feuille,prem_ligne = 1,der_ligne=np.inf):
    ws = wb[nom_feuille]
    tableau = []
    der_ligne = min(der_ligne,ws.max_row)
    for i,row in enumerate(ws.rows,1) : # pour chaque ligne,
        if prem_ligne <= i <= der_ligne :
            ligne = []
            for cell in row : # pour chaque cellule de la ligne,
                ligne.append(cell.value) # lecture de son contenu
            tableau.append(ligne)
    tableau = np.array(tableau) # Conversion en tableau numpy
    shp,typ_don = tableau.shape,tableau.dtype
    return tableau



classeur = "Sujet_N_MutilAnalysis-nettoyé3.1.xlsx"

# creation d'une liste de tous les tableaux contenants les différents sujets/essais
listeFeuilles = [lireFeuille(ouvrirXLSX(classeur)[0],i) for i in (ouvrirXLSX(classeur)[1])]

 #========================
        #Fonction qui va sélectionner les données utiles dans le fichier excel nettoyé
 #========================


#fonction qui retourne des listes de données (TR/Err3D/etc.) en fonction de la Room/du numéro du stimuliSound
def etude_stimuli(L,R,n):
    TR = [] #comme d'habitude, on crée des listes vides que l'on va remplir
    Errh = [] #erreur horizontale
    Errv = [] #erreur verticale
    Errp = [] #erreur profondeur
    Err3D = [] #erreur 3D
    for j in range(len(L)): #pour chaque sujet
        for k in range(1,136): #pour chaque essai
            if int(L[j][k,2]) == R: #si la salle est R
                if int(L[j][k,3]) == n: #ET le son est n
                    if np.isnan(float(L[j][k,8])) == False: #et que en plus la valeur est définie
                        TR.append(float(float(L[j][k,7]))) #alors on ajoute les valeurs correspondantes dans des listes
                        Errh.append(abs(float(L[j][k,9])))
                        Errv.append(abs(float(L[j][k,10])))
                        Errp.append(abs(float(L[j][k,11])))
                        Err3D.append(abs(float(L[j][k,8])))
                   
    print("taille de Err3D =",len(Err3D)) #juste une petite vérification
    return TR,Errh,Err3D #pour alléger le programme, la fonction ne retourne que 3 listes. Pour ploter la donnée que vous voulez, il suffit de modifier "Err3D" par ce que vous voulez
#exemple : return TR,Errh,Err3D devient => return TR,Errh,Errv : cela plotera des graphiques de l'erreur verticale. Il suffit alors de changer les titres du graphique plot et le tour est joué !!
                    


 #========================
        #Création des listes d'erreur 3D correspant à congruent/incongruent dans plage puis dans église puis .. grâce à la fonction précédente
 #========================
  
 #ATTENTION ! toutes cette partie n'utilise que des variables de type Err_3Dk mais cela n'a aucune importance, si la fonction etude_stimuli retourne en 3eme position le TR, alors Err_3Dk sera en fait composé des TR, le nom n'importa pas ici !
  #Ne pas se faire avoir par le nom des variables

#Son de 1 à 9 sont congruents dans Plage_8
Err_3D = []
Err_h = []
for i in range (1,10):
    _,_,Err3D = etude_stimuli(listeFeuilles,8,i) #dans la plage, les sons congruent vont de 1 à 9 donc [1,10[
    for j in range(len(Err3D)):
        Err_3D.append(Err3D[j]) 

print("la taille de Err_3D est : ",len(Err_3D)) #tous les print ici ne servent car vérifier la taille des listes Err3D, ils peuvent totalement être supprimé

    
#Son de 10 à 18 sont congruents dans église_10

Err_3D2 = []
for i in range (10,19):
    _,_,Err3D = etude_stimuli(listeFeuilles,10,i)
    for j in range(len(Err3D)):
        Err_3D2.append(Err3D[j])

print("la taille de Err_3D2 est : ",len(Err_3D2))

#Son de 1 à 9 sont incongruents dans église_10

Err_3D3 = []
for i in range (1,10):
    _,_,Err3D = etude_stimuli(listeFeuilles,10,i)
    for j in range(len(Err3D)):
        Err_3D3.append(Err3D[j])

print("la taille de Err_3D3 est : ",len(Err_3D3))


#Son de 10 à 18 sont incongruents dans plage_8

Err_3D4 = []
for i in range (10,19):
    TR,Errh,Err3D = etude_stimuli(listeFeuilles,8,i)
    for j in range(len(Err3D)):
        Err_3D4.append(Err3D[j])

print("la taille de Err_3D4 est : ",len(Err_3D4))

#Son de 19 à 21 sont neutre dans plage_8 et église_10

Err_3D5 = []
for i in range (19,22):
    TR,Errh,Err3D = etude_stimuli(listeFeuilles,8,i)
    for j in range(len(Err3D)):
        Err_3D5.append(Err3D[j])

print("la taille de Err_3D5 est : ",len(Err_3D5))

Err_3D6 = []
for i in range (19,22):
    TR,Errh,Err3D = etude_stimuli(listeFeuilles,10,i)
    for j in range(len(Err3D)):
        Err_3D6.append(Err3D[j])

print("la taille de Err_3D6 est : ",len(Err_3D6))


Err_3D7 = Err_3D + Err_3D2 #congruents
print("la taille de Err_3D7 est : ",len(Err_3D7))
Err_3D8 = Err_3D3 + Err_3D4 #incongruents
print("la taille de Err_3D8 est : ",len(Err_3D8))
Err_3D9 = Err_3D5 + Err_3D6 #neutres
print("la taille de Err_3D9 est : ",len(Err_3D9))

Err_3D10 = []
for i in range (1,22):
    TR,Errh,Err3D = etude_stimuli(listeFeuilles,0,i)
    for j in range(len(Err3D)):
        Err_3D10.append(Err3D[j])



        


 #========================
        #Plot des listes précédentes
 #========================

### BoxPlot de congruent/incongruent dans plage_8 ###
plt.figure(1)
boxplotElements = plt.boxplot([Err_3D,Err_3D4,Err_3D5], sym = 'b.', whis = 1.5,
                                 widths = [1,1,1], positions = [1,2,3],
                                 patch_artist = True)
plt.gca().xaxis.set_ticklabels(['congruents', 'incongruents','neutres'])
for element in boxplotElements['medians']:
    element.set_color('blue')
    element.set_linewidth(1)
for element in boxplotElements['boxes']:
    element.set_edgecolor('magenta')
    element.set_facecolor('yellow')
    element.set_linewidth(1)
    element.set_linestyle('dashed')
    element.set_fill(True)
    element.set_hatch('/')
for element in boxplotElements['whiskers']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['caps']:
    element.set_color('red')
plt.title('congruent/incongruent et neutre dans plage_8')
plt.ylabel(' TR (s) ')
plt.grid(True)

### BoxPlot de congruent/incongruent dans église_10 ###

plt.figure(2)
boxplotElements = plt.boxplot([Err_3D2,Err_3D3,Err_3D6], sym = 'b.', whis = 1.2,
                                 widths = [1,1,1], positions = [1,2,3],
                                 patch_artist = True)
plt.gca().xaxis.set_ticklabels(['congruents', 'incongruents','neutres'])
for element in boxplotElements['medians']:
    element.set_color('blue')
    element.set_linewidth(1)
for element in boxplotElements['boxes']:
    element.set_edgecolor('magenta')
    element.set_facecolor('yellow')
    element.set_linewidth(1)
    element.set_linestyle('dashed')
    element.set_fill(True)
    element.set_hatch('/')
for element in boxplotElements['whiskers']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['caps']:
    element.set_color('red')
plt.title('congruent/incongruent et neutre dans église_10')
plt.ylabel(' TR (s) ')
plt.grid(True)

### BoxPlot de congruents/incongruents et neutre dans plage/église ###

plt.figure(3)
boxplotElements = plt.boxplot([Err_3D7], sym = 'b.', whis = 1.0, #ploter les boxplot 1 par 1 comme ci-dessous permet d'avoir la main sur la couleur de chaque box, mais cela est très lourds en lignes de codes donc je ne le fais pas à chaque fois
                                 widths = [0.5], positions = [1],
                                 patch_artist = True)
for element in boxplotElements['medians']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['boxes']:
    element.set_edgecolor('black')
    element.set_facecolor('yellow')
    element.set_linewidth(1)
    element.set_linestyle('dashed')
    element.set_fill(True)
    element.set_hatch('/')
for element in boxplotElements['whiskers']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['caps']:
    element.set_color('red')

boxplotElements = plt.boxplot([Err_3D8], sym = 'b.', whis = 1.0,
                                 widths = [0.5], positions = [2],
                                 patch_artist = True)
for element in boxplotElements['medians']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['boxes']:
    element.set_edgecolor('black')
    element.set_facecolor('blue')
    element.set_linewidth(1)
    element.set_linestyle('dashed')
    element.set_fill(True)
    element.set_hatch('/')
for element in boxplotElements['whiskers']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['caps']:
    element.set_color('red')

boxplotElements = plt.boxplot([Err_3D9], sym = 'b.', whis = 1.0,
                                 widths = [0.5], positions = [3],
                                 patch_artist = True)
for element in boxplotElements['medians']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['boxes']:
    element.set_edgecolor('black')
    element.set_facecolor('green')
    element.set_linewidth(1)
    element.set_linestyle('dashed')
    element.set_fill(True)
    element.set_hatch('/')
for element in boxplotElements['whiskers']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['caps']:
    element.set_color('red')

boxplotElements = plt.boxplot([Err_3D10], sym = 'b.', whis = 1.0,
                                 widths = [0.5], positions = [4],
                                 patch_artist = True)

plt.gca().xaxis.set_ticklabels(['congruents', 'incongruents','neutres','non visual'])
for element in boxplotElements['medians']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['boxes']:
    element.set_edgecolor('black')
    element.set_facecolor('magenta')
    element.set_linewidth(1)
    element.set_linestyle('dashed')
    element.set_fill(True)
    element.set_hatch('/')
for element in boxplotElements['whiskers']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['caps']:
    element.set_color('red')
    
plt.title(' Comparaison TR') #il faut changer les titres pour s'adapter à ce que retourne la fonction etude_stimuli (ligne 68)
plt.ylabel(' TR (s) ')
plt.grid(True)

plt.show()



