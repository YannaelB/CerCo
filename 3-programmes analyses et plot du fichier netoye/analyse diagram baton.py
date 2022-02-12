import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
import matplotlib.pyplot as plt
import os
from openpyxl.styles import Color,PatternFill
from numpy.linalg import solve
from collections import Counter
import statistics as st
from scipy import stats


def ouvrirXLSX(filepath) :
    wb = load_workbook(filepath)
    return wb,wb.sheetnames

def lireFeuille(wb,nom_feuille,prem_ligne = 1,der_ligne=np.inf):
    ws = wb[nom_feuille]
    tableau = []
    der_ligne = min(der_ligne,ws.max_row)
    for i,row in enumerate(ws.rows,1) : # pour chaque ligne,
        if prem_ligne <= i <= der_ligne :
            ligne = []
            for cell in row : # pour chaque cellule de la ligne,
                ligne.append(cell.value) # lecture de son contenu
            tableau.append(ligne)
    tableau = np.array(tableau) # Conversion en tableau numpy
    shp,typ_don = tableau.shape,tableau.dtype
    return tableau



classeur = "Sujet_N_MutilAnalysis-nettoyé3.1.xlsx"


listeFeuilles = [lireFeuille(ouvrirXLSX(classeur)[0],i) for i in (ouvrirXLSX(classeur)[1])]




def etude_stimuli(L,n):
    TR = []
    Errh = []
    Err3D = []
    for j in range(len(L)):
        print('sujet numéro : ',j)
        for k in range(1,136):
            if int(L[j][k,3]) == n:
                if np.isnan(float(L[j][k,8])) == False:
                    TR.append(float(float(L[j][k,7])))
                    Errh.append(abs(float(L[j][k,9])))
                    Err3D.append(abs(float(L[j][k,8])))
    print("son numéro :",n)
    print("taille de TR =",len(TR))
    print("taille de Errh =",len(Errh))
    print("taille de Err3D =",len(Err3D))
    return TR,st.mean(TR),st.pstdev(TR),Errh,st.mean(Errh),st.pstdev(Errh),Err3D,st.mean(Err3D)*100,st.pstdev(Err3D)*100
                    

def calcul_stat_congruent(L,a,b):
    Tps_react = []
    Ectype_TR = []
    err_h = []
    Ectype_Errh = []
    err_3D = []
    Ectype_Err3D = []

    for i in range(a,b+1):
        print("son numéro : ",i)
        TR,M_TR,Ec_TR,Errh,M_Errh,Ec_Errh,Err3D,M_Err3D,Ec_Err3D = etude_stimuli(L,i)
        Tps_react.append(M_TR)
        Ectype_TR.append(Ec_TR)
        err_h.append(M_Errh)
        Ectype_Errh.append(Ec_Errh)
        err_3D.append(M_Err3D)
        Ectype_Err3D.append(Ec_Err3D)

    print("Voici Tps_react :",Tps_react)                     


    moy_TR = st.mean(Tps_react)
    ecty_TR = st.pstdev(Tps_react)
    moy_ecarttype_TR = st.mean(Ectype_TR)
    
    moy_Errh = st.mean(err_h)
    ecty_Errh = st.pstdev(err_h)
    moy_ecarttype_Errh = st.mean(Ectype_Errh)
    
    moy_Err3D = st.mean(err_3D)
    ecty_Err3D = st.pstdev(err_3D)
    moy_ecarttype_Err3D = st.mean(Ectype_Err3D)

    return moy_TR,moy_ecarttype_TR,moy_Errh,moy_ecarttype_Errh,moy_Err3D,moy_ecarttype_Err3D
    
                        

moy_TR,moy_ecarttype_TR,moy_Errh,moy_ecarttype_Errh,moy_Err3D,moy_ecarttype_Err3D = calcul_stat_congruent(listeFeuilles,1,9)
moy_TR2,moy_ecarttype_TR2,moy_Errh2,moy_ecarttype_Errh2,moy_Err3D2,moy_ecarttype_Err3D2 = calcul_stat_congruent(listeFeuilles,10,18)
moy_TR3,moy_ecarttype_TR3,moy_Errh3,moy_ecarttype_Errh3,moy_Err3D3,moy_ecarttype_Err3D3 = calcul_stat_congruent(listeFeuilles,19,21)



x = ["congruent","incongruent","neutre"]
y1 = [moy_Err3D,moy_Err3D2,moy_Err3D3]
y2 = [moy_Err3D-moy_ecarttype_Err3D,moy_Err3D2-moy_ecarttype_Err3D2,moy_Err3D3-moy_ecarttype_Err3D3]
y3 = [moy_Err3D+moy_ecarttype_Err3D,moy_Err3D2+moy_ecarttype_Err3D2,moy_Err3D3+moy_ecarttype_Err3D3]
y4 = [moy_ecarttype_Err3D,moy_ecarttype_Err3D2,moy_ecarttype_Err3D3]

plt.figure(1)
plt.bar(x,y1, width = 0.99, color = 'green')
#plt.errorbar(x, y1, [y2,y3],fmt='.k', ecolor='gray', lw=1)
plt.errorbar(x, y1, y4,fmt='.k',ecolor='gray', lw=2)
plt.xlabel(" contexte ")
plt.ylabel(' Err azimuth (°) ')
plt.title(" erreur azimuth localisation son selon contexte ")
#plt.ylim(110,200)
#plt.grid(True)
#plt.show()

    

plt.show()





