import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
import matplotlib.pyplot as plt
import os
from openpyxl.styles import Color,PatternFill
from numpy.linalg import solve
from collections import Counter
import statistics as st
from scipy import stats


def ouvrirXLSX(filepath) :
    wb = load_workbook(filepath)
    return wb,wb.sheetnames

def lireFeuille(wb,nom_feuille,prem_ligne = 1,der_ligne=np.inf):
    ws = wb[nom_feuille]
    tableau = []
    der_ligne = min(der_ligne,ws.max_row)
    for i,row in enumerate(ws.rows,1) : # pour chaque ligne,
        if prem_ligne <= i <= der_ligne :
            ligne = []
            for cell in row : # pour chaque cellule de la ligne,
                ligne.append(cell.value) # lecture de son contenu
            tableau.append(ligne)
    tableau = np.array(tableau) # Conversion en tableau numpy
    shp,typ_don = tableau.shape,tableau.dtype
    return tableau



classeur = "Sujet_N_MutilAnalysis-nettoyé3.1.xlsx"


listeFeuilles = [lireFeuille(ouvrirXLSX(classeur)[0],i) for i in (ouvrirXLSX(classeur)[1])]




def etude_stimuli(L,R,n):
    TR = []
    Errh = []
    Errv = []
    Errp = []
    Err3D = []
    ErrHead = []
    for j in range(len(L)):
        #print('sujet numéro : ',j)
        for k in range(1,136):
            if int(L[j][k,2]) == R:
                if int(L[j][k,3]) == n:
                    if np.isnan(float(L[j][k,8])) == False:
                        TR.append(float(float(L[j][k,7])))
                        Errh.append(abs(float(L[j][k,9])))
                        Errv.append(abs(float(L[j][k,10])))
                        Errp.append(abs(float(L[j][k,11])))
                        Err3D.append(abs(float(L[j][k,8])))
                    if np.isnan(float(L[j][k,12])) == False:
                        ErrHead.append(abs(float(L[j][k,12])))
                   
    print("son numéro :",n," dans scène numéro :", R)
    print("taille de TR =",len(TR))
    print("taille de Errh =",len(Errh))
    print("taille de Err3D =",len(Err3D))
    return TR,Errh,ErrHead
                    


    
                        
#Son de 1 à 9 sont congruents dans Plage_8
Err_3D = []
Err_h = []
for i in range (1,10):
    TR,Errh,Err3D = etude_stimuli(listeFeuilles,8,i)
    for j in range(len(Err3D)):
        Err_3D.append(Err3D[j])

print("la taille de Err_3D est : ",len(Err_3D))

    
#Son de 10 à 18 sont congruents dans église_10

Err_3D2 = []
for i in range (10,19):
    TR,Errh,Err3D = etude_stimuli(listeFeuilles,10,i)
    for j in range(len(Err3D)):
        Err_3D2.append(Err3D[j])

print("la taille de Err_3D2 est : ",len(Err_3D2))

#Son de 1 à 9 sont incongruents dans église_10

Err_3D3 = []
for i in range (1,10):
    TR,Errh,Err3D = etude_stimuli(listeFeuilles,10,i)
    for j in range(len(Err3D)):
        Err_3D3.append(Err3D[j])

print("la taille de Err_3D3 est : ",len(Err_3D3))


#Son de 10 à 18 sont incongruents dans plage_8

Err_3D4 = []
for i in range (10,19):
    TR,Errh,Err3D = etude_stimuli(listeFeuilles,8,i)
    for j in range(len(Err3D)):
        Err_3D4.append(Err3D[j])

print("la taille de Err_3D4 est : ",len(Err_3D4))

#Son de 19 à 21 sont neutre dans plage_8 et église_10

Err_3D5 = []
for i in range (19,22):
    TR,Errh,Err3D = etude_stimuli(listeFeuilles,8,i)
    for j in range(len(Err3D)):
        Err_3D5.append(Err3D[j])

print("la taille de Err_3D5 est : ",len(Err_3D5))

Err_3D6 = []
for i in range (19,22):
    TR,Errh,Err3D = etude_stimuli(listeFeuilles,10,i)
    for j in range(len(Err3D)):
        Err_3D6.append(Err3D[j])

print("la taille de Err_3D6 est : ",len(Err_3D6))


Err_3D7 = Err_3D + Err_3D2 #congruents
print("la taille de Err_3D7 est : ",len(Err_3D7))
Err_3D8 = Err_3D3 + Err_3D4 #incongruents
print("la taille de Err_3D8 est : ",len(Err_3D8))
Err_3D9 = Err_3D5 + Err_3D6 #neutres
print("la taille de Err_3D9 est : ",len(Err_3D9))

Err_3D10 = []
for i in range (1,22):
    TR,Errh,Err3D = etude_stimuli(listeFeuilles,0,i)
    for j in range(len(Err3D)):
        Err_3D10.append(Err3D[j])



        
### BoxPlot de congruent/incongruent dans plage_8 ###

plt.figure(1)
boxplotElements = plt.boxplot([Err_3D,Err_3D4,Err_3D5], sym = 'b.', whis = 1.5,
                                 widths = [1,1,1], positions = [1,2,3],
                                 patch_artist = True)
plt.gca().xaxis.set_ticklabels(['congruents', 'incongruents','neutres'])
for element in boxplotElements['medians']:
    element.set_color('blue')
    element.set_linewidth(1)
for element in boxplotElements['boxes']:
    element.set_edgecolor('magenta')
    element.set_facecolor('yellow')
    element.set_linewidth(1)
    element.set_linestyle('dashed')
    element.set_fill(True)
    element.set_hatch('/')
for element in boxplotElements['whiskers']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['caps']:
    element.set_color('red')
plt.title('congruent/incongruent et neutre dans plage_8')
plt.ylabel(' erreur azimuth (°) ')
plt.grid(True)

### BoxPlot de congruent/incongruent dans église_10 ###

plt.figure(2)
boxplotElements = plt.boxplot([Err_3D2,Err_3D3,Err_3D6], sym = 'b.', whis = 1.2,
                                 widths = [1,1,1], positions = [1,2,3],
                                 patch_artist = True)
plt.gca().xaxis.set_ticklabels(['congruents', 'incongruents','neutres'])
for element in boxplotElements['medians']:
    element.set_color('blue')
    element.set_linewidth(1)
for element in boxplotElements['boxes']:
    element.set_edgecolor('magenta')
    element.set_facecolor('yellow')
    element.set_linewidth(1)
    element.set_linestyle('dashed')
    element.set_fill(True)
    element.set_hatch('/')
for element in boxplotElements['whiskers']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['caps']:
    element.set_color('red')
plt.title('congruent/incongruent et neutre dans église_10')
plt.ylabel(' erreur azimuth tête (°) ')
plt.grid(True)

### BoxPlot de congruents/incongruents et neutre dans plage/église ###

plt.figure(3)
boxplotElements = plt.boxplot([Err_3D7], sym = 'b.', whis = 1.0,
                                 widths = [0.5], positions = [1],
                                 patch_artist = True)
for element in boxplotElements['medians']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['boxes']:
    element.set_edgecolor('black')
    element.set_facecolor('yellow')
    element.set_linewidth(1)
    element.set_linestyle('dashed')
    element.set_fill(True)
    element.set_hatch('/')
for element in boxplotElements['whiskers']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['caps']:
    element.set_color('red')

boxplotElements = plt.boxplot([Err_3D8], sym = 'b.', whis = 1.0,
                                 widths = [0.5], positions = [2],
                                 patch_artist = True)
for element in boxplotElements['medians']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['boxes']:
    element.set_edgecolor('black')
    element.set_facecolor('blue')
    element.set_linewidth(1)
    element.set_linestyle('dashed')
    element.set_fill(True)
    element.set_hatch('/')
for element in boxplotElements['whiskers']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['caps']:
    element.set_color('red')

boxplotElements = plt.boxplot([Err_3D9], sym = 'b.', whis = 1.0,
                                 widths = [0.5], positions = [3],
                                 patch_artist = True)
for element in boxplotElements['medians']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['boxes']:
    element.set_edgecolor('black')
    element.set_facecolor('green')
    element.set_linewidth(1)
    element.set_linestyle('dashed')
    element.set_fill(True)
    element.set_hatch('/')
for element in boxplotElements['whiskers']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['caps']:
    element.set_color('red')

boxplotElements = plt.boxplot([Err_3D10], sym = 'b.', whis = 1.0,
                                 widths = [0.5], positions = [4],
                                 patch_artist = True)

plt.gca().xaxis.set_ticklabels(['congruents', 'incongruents','neutres','non visual'])
for element in boxplotElements['medians']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['boxes']:
    element.set_edgecolor('black')
    element.set_facecolor('magenta')
    element.set_linewidth(1)
    element.set_linestyle('dashed')
    element.set_fill(True)
    element.set_hatch('/')
for element in boxplotElements['whiskers']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['caps']:
    element.set_color('red')
    
plt.title(' Comparaison erreur azimuth tête')
plt.ylabel(' erreur azimuth tête (°) ')
plt.grid(True)

plt.show()



