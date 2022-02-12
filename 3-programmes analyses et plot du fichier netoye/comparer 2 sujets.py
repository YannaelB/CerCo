import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
import matplotlib.pyplot as plt
import os
from openpyxl.styles import Color,PatternFill
from numpy.linalg import solve
from collections import Counter
import statistics as st
from scipy import stats


def ouvrirXLSX(filepath) :
    wb = load_workbook(filepath)
    return wb,wb.sheetnames

def lireFeuille(wb,nom_feuille,prem_ligne = 1,der_ligne=np.inf):
    ws = wb[nom_feuille]
    tableau = []
    der_ligne = min(der_ligne,ws.max_row)
    for i,row in enumerate(ws.rows,1) : # pour chaque ligne,
        if prem_ligne <= i <= der_ligne :
            ligne = []
            for cell in row : # pour chaque cellule de la ligne,
                ligne.append(cell.value) # lecture de son contenu
            tableau.append(ligne)
    tableau = np.array(tableau) # Conversion en tableau numpy
    shp,typ_don = tableau.shape,tableau.dtype
    return tableau



classeur = "Sujet_N_MutilAnalysis-nettoyé3.1.xlsx"


listeFeuilles = [lireFeuille(ouvrirXLSX(classeur)[0],i) for i in (ouvrirXLSX(classeur)[1])]


def etude_stimuli(L,L_s,R,n):
    nb_hand = []
    nb_head = []
    TR = []
    Errh = []
    Errv = []
    Errp = []
    Err3D = []
    ErrHead = []
    for s in L_s:
        for k in range(1,136):
            if int(L[s][k,2]) == R:
                if int(L[s][k,3]) == n: 
                    if np.isnan(float(L[s][k,8])) == False:
                        TR.append(float(float(L[s][k,7])))
                        nb_hand.append(float(float(L[s][k,6])))
                        nb_head.append(float(float(L[s][k,5])))
                        Errh.append(abs(float(L[s][k,9])))
                        Errv.append(abs(float(L[s][k,10])))
                        Errp.append(abs(float(L[s][k,11])))
                        Err3D.append(abs(float(L[s][k,8])))
                        ErrHead.append(abs(float(L[s][k,12])))
                        

    return nb_head,nb_hand,TR,Err3D,Errh,Errv,Errp,ErrHead


def etude_comparative(L,Ls1,Ls2,n):

    Err_3D = []
    Err_3D2 = []
    Err = []
    Err2 = []
    for i in (8,10,0):

        print("/// DANS LA SALLE : ", i," // pour le son numéro : ",n," /////")
        nb_head,nb_hand,TR,Err3D,Errh,Errv,Errp,ErrHead = etude_stimuli(L,Ls1,i,n)
        nb_head2,nb_hand2,TR2,Err3D2,Errh2,Errv2,Errp2,ErrHead2 = etude_stimuli(L,Ls2,i,n)
        Err_3D.append(Err3D)
        Err_3D2.append(Err3D2)

        print("taille de TR est :",len(TR)," la taille de TR2 est :",len(TR2))
        print("taille de Err3D est :",len(Err3D)," la taille de Err3D2 est :",len(Err3D2))

        print("moyenne Err3D: ",st.mean(Err3D),"moyenne Err3D2:",st.mean(Err3D2))
        ttest1= stats.ttest_ind(Err3D, Err3D2, equal_var = False)
        print('t.test sur erreur 3D  : ', ttest1)
        print("  ")
        print("moyenne Errh: ",st.mean(Errh),"moyenne Errh2:",st.mean(Errh2))
        ttest2= stats.ttest_ind(Errh, Errh2, equal_var = False)
        print('t.test sur erreur h  : ', ttest2)
        print("  ")
        print("moyenne Errv: ",st.mean(Errv),"moyenne Errv2:",st.mean(Errv2))
        ttest3= stats.ttest_ind(Errv, Errv2, equal_var = False)
        print('t.test sur erreur v  : ', ttest3)
        print("  ")
        print("moyenne Errp: ",st.mean(Errp),"moyenne Errp2:",st.mean(Errp2))
        ttest4= stats.ttest_ind(Errp, Errp2, equal_var = False)
        print('t.test sur erreur p  : ', ttest4)
        print("  ")
        print("moyenne ErrHead: ",st.mean(ErrHead),"moyenne ErrHead2:",st.mean(ErrHead2))
        ttest5= stats.ttest_ind(ErrHead, ErrHead2, equal_var = False)
        print('t.test sur erreur h  : ', ttest5)
        print("  ")

    for j in range(len(Err_3D)):
        for k in range(len(Err_3D[j])):
            Err.append(Err_3D[j][k])
            
    for j in range(len(Err_3D2)):
        for k in range(len(Err_3D2[j])):
            Err2.append(Err_3D2[j][k])

    print("moyenne Err: ",st.mean(Err),"moyenne Err2:",st.mean(Err2))
    ttest6= stats.ttest_ind(Err, Err2, equal_var = False)
    print('t.test sur erreur 3D global  : ', ttest6)



    print("1: ",0.2019965176206825+0.2184372102294057+0.2038628505960198)
    print("2: ",0.19926674661497523+0.20043945188615525+0.23695450435324836)
        

        
#utilisateur well identified bateau :
        
etude_comparative(listeFeuilles,(0,4,5,7,9,11,13),(2,3,6,8,10,12),2)  #bateau

etude_comparative(listeFeuilles,(1,4,5,7,8,9,13),(0,2,3,6,10,11,12),1)  #avion

etude_comparative(listeFeuilles,(1,4,5,6,10,12),(0,2,3,7,8,9,11,13),20)  #neutre2





#nb_head1,nb_hand1,TR1,Err3D1,Errh1,Errv1,Errp1 = etude_stimuli(listeFeuilles,(0,4,5),8,2)
#nb_head2,nb_hand2,TR1,Err3D2,Errh2,Errv2,Errp2 = etude_stimuli(listeFeuilles,(0,4,5),10,2)
#nb_head3,nb_hand3,TR1,Err3D3,Errh3,Errv3,Errp3 = etude_stimuli(listeFeuilles,(2,3,6),8,2)
#nb_head4,nb_hand4,TR1,Err3D4,Errh4,Errv4,Errp4 = etude_stimuli(listeFeuilles,(2,3,6),10,2)




plt.figure(1)
boxplotElements = plt.boxplot(Errv1, sym = 'b.', whis = 1.0,
                                 widths = [0.5], positions = [1],
                                 patch_artist = True)
for element in boxplotElements['medians']:
    element.set_color('black')
    element.set_linewidth(1)
for element in boxplotElements['boxes']:
    element.set_edgecolor('black')
    element.set_facecolor('yellow')
    element.set_linewidth(1)
    element.set_linestyle('dashed')
    element.set_fill(True)
    element.set_hatch('/')
for element in boxplotElements['whiskers']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['caps']:
    element.set_color('red')

boxplotElements = plt.boxplot(Errv3, sym = 'b.', whis = 1.0,
                                 widths = [0.5], positions = [2],
                                 patch_artist = True)

plt.gca().xaxis.set_ticklabels(['Beach', 'not Beach'])
for element in boxplotElements['medians']:
    element.set_color('black')
    element.set_linewidth(1)
for element in boxplotElements['boxes']:
    element.set_edgecolor('black')
    element.set_facecolor('magenta')
    element.set_linewidth(1)
    element.set_linestyle('dashed')
    element.set_fill(True)
    element.set_hatch('/')
for element in boxplotElements['whiskers']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['caps']:
    element.set_color('red')
    
plt.title(' Comparaison Erreur Elevation dans la scene de plage')
plt.ylabel(' ErrV (°) ')
plt.grid(True)

#plt.show()
