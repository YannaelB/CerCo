import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
import matplotlib.pyplot as plt
import os
from openpyxl.styles import Color,PatternFill
from numpy.linalg import solve
from collections import Counter
import statistics as st
from scipy import stats


def ouvrirXLSX(filepath) :
    wb = load_workbook(filepath)
    return wb,wb.sheetnames

def lireFeuille(wb,nom_feuille,prem_ligne = 1,der_ligne=np.inf):
    ws = wb[nom_feuille]
    tableau = []
    der_ligne = min(der_ligne,ws.max_row)
    for i,row in enumerate(ws.rows,1) : # pour chaque ligne,
        if prem_ligne <= i <= der_ligne :
            ligne = []
            for cell in row : # pour chaque cellule de la ligne,
                ligne.append(cell.value) # lecture de son contenu
            tableau.append(ligne)
    tableau = np.array(tableau) # Conversion en tableau numpy
    shp,typ_don = tableau.shape,tableau.dtype
    return tableau



classeur = "Sujet_N_MutilAnalysis-nettoyé3.1.xlsx"


listeFeuilles = [lireFeuille(ouvrirXLSX(classeur)[0],i) for i in (ouvrirXLSX(classeur)[1])]


def etude_stimuli(L,R,n):
    nb_hand = []
    nb_head = []
    TR = []
    Errh = []
    Errv = []
    Errp = []
    Err3D = []
    for s in range(len(L)):
        for k in range(1,136):
            if int(L[s][k,2]) == R:
                if int(L[s][k,3]) == n: 
                    if np.isnan(float(L[s][k,8])) == False:
                        TR.append(float(float(L[s][k,7])))
                        nb_hand.append(float(float(L[s][k,6])))
                        nb_head.append(float(float(L[s][k,5])))
                        Errh.append(abs(float(L[s][k,9])))
                        Errv.append(abs(float(L[s][k,10])))
                        Errp.append(abs(float(L[s][k,11])))
                        Err3D.append(abs(float(L[s][k,8])))

    return nb_head,nb_hand,TR,Err3D,Errh,Errv,Errp

def in_cong(L,R,n1,n2):
    nb_hand1 = []
    nb_head1 = []
    TR1 = []
    Errh1 = []
    Errv1 = []
    Errp1 = []
    Err3D1 = []
    for i in range(n1,n2+1):
        nb_head,nb_hand,TR,Err3D,Errh,Errv,Errp = etude_stimuli(L,R,i)
        for j in range(len(Errh)):
            nb_head1.append(nb_head[j])
            nb_hand1.append(nb_hand[j])
            TR1.append(TR[j])
            Errh1.append(Errh[j])
            Errv1.append(Errv[j])
            Errp1.append(Errp[j])
            Err3D1.append(Err3D[j])
    return Errv1,Errp1
            

    

a1,b1 =in_cong(listeFeuilles,10,1,9)
a2,b2 =in_cong(listeFeuilles,10,10,18)
a3,b3 =in_cong(listeFeuilles,10,19,21)

a1,b1 =in_cong(listeFeuilles,0,1,9)
a2,b2 =in_cong(listeFeuilles,0,10,18)
a3,b3 =in_cong(listeFeuilles,0,19,21)                
                



plt.figure(1)

boxplotElements = plt.boxplot(a1, sym = 'b.', whis = 1.0,
                                 widths = [0.5], positions = [2],
                                 patch_artist = True)
for element in boxplotElements['medians']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['boxes']:
    element.set_edgecolor('black')
    element.set_facecolor('blue')
    element.set_linewidth(1)
    element.set_linestyle('dashed')
    element.set_fill(True)
    element.set_hatch('/')
for element in boxplotElements['whiskers']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['caps']:
    element.set_color('red')

boxplotElements = plt.boxplot(a2, sym = 'b.', whis = 1.0,
                                 widths = [0.5], positions = [3],
                                 patch_artist = True)
for element in boxplotElements['medians']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['boxes']:
    element.set_edgecolor('black')
    element.set_facecolor('green')
    element.set_linewidth(1)
    element.set_linestyle('dashed')
    element.set_fill(True)
    element.set_hatch('/')
for element in boxplotElements['whiskers']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['caps']:
    element.set_color('red')

boxplotElements = plt.boxplot(a3, sym = 'b.', whis = 1.0,
                                 widths = [0.5], positions = [4],
                                 patch_artist = True)

plt.gca().xaxis.set_ticklabels(['Beach', 'Church','neutres'])
for element in boxplotElements['medians']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['boxes']:
    element.set_edgecolor('black')
    element.set_facecolor('magenta')
    element.set_linewidth(1)
    element.set_linestyle('dashed')
    element.set_fill(True)
    element.set_hatch('/')
for element in boxplotElements['whiskers']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['caps']:
    element.set_color('red')
    
plt.title(' Comparaison Erreur Elevation dans Church')
plt.ylabel(' ErrV (°) ')
plt.grid(True)

plt.figure(2)

boxplotElements = plt.boxplot(b1, sym = 'b.', whis = 1.0,
                                 widths = [0.5], positions = [2],
                                 patch_artist = True)
for element in boxplotElements['medians']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['boxes']:
    element.set_edgecolor('black')
    element.set_facecolor('blue')
    element.set_linewidth(1)
    element.set_linestyle('dashed')
    element.set_fill(True)
    element.set_hatch('/')
for element in boxplotElements['whiskers']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['caps']:
    element.set_color('red')

boxplotElements = plt.boxplot(b2, sym = 'b.', whis = 1.0,
                                 widths = [0.5], positions = [4],
                                 patch_artist = True)
for element in boxplotElements['medians']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['boxes']:
    element.set_edgecolor('black')
    element.set_facecolor('green')
    element.set_linewidth(1)
    element.set_linestyle('dashed')
    element.set_fill(True)
    element.set_hatch('/')
for element in boxplotElements['whiskers']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['caps']:
    element.set_color('red')

boxplotElements = plt.boxplot(b3, sym = 'b.', whis = 1.0,
                                 widths = [0.5], positions = [3],
                                 patch_artist = True)

plt.gca().xaxis.set_ticklabels(['Beach', 'Church','neutres'])
for element in boxplotElements['medians']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['boxes']:
    element.set_edgecolor('black')
    element.set_facecolor('magenta')
    element.set_linewidth(1)
    element.set_linestyle('dashed')
    element.set_fill(True)
    element.set_hatch('/')
for element in boxplotElements['whiskers']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['caps']:
    element.set_color('red')
    
plt.title(' Comparaison Erreur profondeur dans Gris')
plt.ylabel(' ErrP (m) ')
plt.grid(True)

plt.show()
        

        
        






        
