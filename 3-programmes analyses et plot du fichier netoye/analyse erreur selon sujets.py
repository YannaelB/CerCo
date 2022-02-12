import numpy as np
import matplotlib as mpl
from openpyxl import load_workbook
from openpyxl import Workbook
import matplotlib.pyplot as plt
import os
from openpyxl.styles import Color,PatternFill
from numpy.linalg import solve
from collections import Counter
import statistics as st
from scipy import stats
import matplotlib.ticker as mticker


def ouvrirXLSX(filepath) :
    wb = load_workbook(filepath)
    return wb,wb.sheetnames

def lireFeuille(wb,nom_feuille,prem_ligne = 1,der_ligne=np.inf):
    ws = wb[nom_feuille]
    tableau = []
    der_ligne = min(der_ligne,ws.max_row)
    for i,row in enumerate(ws.rows,1) : # pour chaque ligne,
        if prem_ligne <= i <= der_ligne :
            ligne = []
            for cell in row : # pour chaque cellule de la ligne,
                ligne.append(cell.value) # lecture de son contenu
            tableau.append(ligne)
    tableau = np.array(tableau) # Conversion en tableau numpy
    shp,typ_don = tableau.shape,tableau.dtype
    return tableau



classeur = "Sujet_N_MutilAnalysis-nettoyé2.xlsx"


listeFeuilles = [lireFeuille(ouvrirXLSX(classeur)[0],i) for i in (ouvrirXLSX(classeur)[1])]




def etude_stimuli(L):
    TR = []
    Errh = []
    Errv = []
    Err3D = []
    for j in range(len(L)):
        s_TR = []
        s_Errh = []
        s_Errv = []
        s_Err3D = []
        for k in range(1,136):
            if np.isnan(float(L[j][k,8])) == False:
                s_TR.append(float(float(L[j][k,7])))
                s_Errh.append(abs(float(L[j][k,9])))
                s_Err3D.append(abs(float(L[j][k,8])))
                s_Errv.append(abs(float(L[j][k,10])))
        TR.append(s_TR)
        Errh.append(s_Errh)
        Errv.append(s_Errv)
        Err3D.append(s_Err3D)

    print("taille de TR =",len(TR))
    print("taille de Errh =",len(Errh))
    print(" TR =",TR[0][:5])
    #print(" Errh =", Errh)
    print("taille de Err3D =",len(Err3D))
    return TR,Errh,Errv,Err3D
                    


    

TR,Errh,Errv,Err3D = etude_stimuli(listeFeuilles)
   


for i in range(len(Err3D)):
    print(" la moyenne du sujet : ", i+1,"est :",st.mean(Err3D[i]))




### BoxPlot par sujet ###

plt.figure(1)
boxplotElements = plt.boxplot([Errh[0],Errh[1],Errh[2],Errh[3],Errh[4],Errh[5],Errh[6]], sym = 'b.', whis = 1.5,
                                 widths = [1,1,1,1,1,1,1], positions = [1,3,5,7,9,11,13],
                                 patch_artist = True)
for element in boxplotElements['medians']:
    element.set_color('blue')
    element.set_linewidth(1)
for element in boxplotElements['boxes']:
    element.set_edgecolor('magenta')
    element.set_facecolor('yellow')
    element.set_linewidth(1)
    element.set_linestyle('dashed')
    element.set_fill(True)
    element.set_hatch('/')
for element in boxplotElements['whiskers']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['caps']:
    element.set_color('red')

boxplotElements = plt.boxplot([Errv[0],Errv[1],Errv[2],Errv[3],Errv[4],Errv[5],Errv[6]], sym = 'b.', whis = 1.5,
                                 widths = [1,1,1,1,1,1,1], positions = [2,4,6,8,10,12,14],
                                 patch_artist = True)
plt.gca().xaxis.set_ticklabels(['S1-Errh', 'S1-Errv','S2-Errh', 'S2-Errv', 'S3-Errh', 'S3-Errv','S4-Errh', 'S4-Errv','S5-Errh', 'S5-Errv','S6-Errh', 'S6-Errv','S7-Errh', 'S7-Errv'])
for element in boxplotElements['medians']:
    element.set_color('blue')
    element.set_linewidth(1)
for element in boxplotElements['boxes']:
    element.set_edgecolor('magenta')
    element.set_facecolor('green')
    element.set_linewidth(1)
    element.set_linestyle('dashed')
    element.set_fill(True)
    element.set_hatch('/')
for element in boxplotElements['whiskers']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['caps']:
    element.set_color('red')
plt.title('Erreur de localisation selon les sujets')
plt.ylabel(' Erreur Azimuth/Elevation (°) ')
plt.grid(True)


plt.figure(2)
boxplotElements = plt.boxplot([Err3D[0],Err3D[1],Err3D[2],Err3D[3],Err3D[4],Err3D[5],Err3D[6]], sym = 'b.', whis = 1.5,
                                 widths = [0.6,0.6,0.6,0.6,0.6,0.6,0.6], positions = [1,2,3,4,5,6,7],
                                 patch_artist = True)
plt.gca().xaxis.set_ticklabels(['Sujet1', 'Sujet2','Sujet3','Sujet4','Sujet5','Sujet6','Sujet7'])
for element in boxplotElements['medians']:
    element.set_color('blue')
    element.set_linewidth(1)
for element in boxplotElements['boxes']:
    element.set_edgecolor('black')
    element.set_facecolor('magenta')
    element.set_linewidth(1)
    element.set_linestyle('dashed')
    element.set_fill(True)
    element.set_hatch('/')
for element in boxplotElements['whiskers']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['caps']:
    element.set_color('red')
plt.title('Erreur 3D selon les sujets')
plt.ylabel(' Erreur 3D (m) ')
plt.grid(True)

plt.figure(3)
boxplotElements = plt.boxplot([TR[0],TR[1],TR[2],TR[3],TR[4],TR[5],TR[6]], sym = 'b.', whis = 1.5,
                                 widths = [1,1,1,1,1,1,1], positions = [1,2,3,4,5,6,7],
                                 patch_artist = True)
plt.gca().xaxis.set_ticklabels(['S1_TR', 'S2_TR','S3_TR','S4_TR','S5_TR','S6_TR','S7_TR'])
for element in boxplotElements['medians']:
    element.set_color('blue')
    element.set_linewidth(1)
for element in boxplotElements['boxes']:
    element.set_edgecolor('magenta')
    element.set_facecolor('yellow')
    element.set_linewidth(1)
    element.set_linestyle('dashed')
    element.set_fill(True)
    element.set_hatch('/')
for element in boxplotElements['whiskers']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['caps']:
    element.set_color('red')
plt.title('TR selon les sujets')
plt.ylabel(' TR (s) ')
plt.grid(True)

#positions = [1,2,3,4,5,6,7,8,9,10,11,12,13,14],

plt.figure(4)
plt.violinplot([Err3D[0],Err3D[1],Err3D[2],Err3D[3],Err3D[4],Err3D[5],Err3D[6],Err3D[7],Err3D[8],Err3D[9],Err3D[10],Err3D[11],Err3D[12],Err3D[13]],points=300, widths=0.8, 
                     showmeans=True, showextrema=True, showmedians=True)

label_format = '{:,.0f}'
positions = [1, 2, 3, 4, 5,6,7,8,9,10,11,12,13,14]
labels = ['S1', 'S2','S3','S4','S5','S6','S7','S8', 'S9','S10','S11','S12','S13','S14']
#plt.gca().xaxis.set_ticklabels(['S1', 'S2','S3','S4','S5','S6','S7','S8', 'S9','S10','S11','S12','S13','S14'])
#label_format = '{:.1%}'
plt.gca().xaxis.set_major_locator(mticker.FixedLocator(positions))
plt.gca().xaxis.set_major_formatter(mticker.FixedFormatter(labels))
plt.grid()
plt.title('Erreur 3D selon les sujets')
plt.ylabel(' Erreur 3D (m) ')


plt.figure(6)
boxplotElements = plt.boxplot([Err3D[0],Err3D[1],Err3D[2],Err3D[3],Err3D[4],Err3D[5],Err3D[6],Err3D[7],Err3D[8],Err3D[9],Err3D[10],Err3D[11],Err3D[12],Err3D[13]], sym = 'b.', whis = 1,
                                 widths = [0.6,0.6,0.6,0.6,0.6,0.6,0.6,0.6,0.6,0.6,0.6,0.6,0.6,0.6], positions = [1,2,3,4,5,6,7,8,9,10,11,12,13,14],
                                 patch_artist = True)
plt.gca().xaxis.set_ticklabels(['S1','S2','S3','S4','S5','S6','S7','S8','S9','S10','S11','S12','S13','S14'])
for element in boxplotElements['medians']:
    element.set_color('blue')
    element.set_linewidth(1)
for element in boxplotElements['boxes']:
    element.set_edgecolor('black')
    element.set_facecolor('magenta')
    element.set_linewidth(1)
    element.set_linestyle('dashed')
    element.set_fill(True)
    element.set_hatch('/')
for element in boxplotElements['whiskers']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['caps']:
    element.set_color('red')
plt.title('Erreur 3D selon les sujets')
plt.ylabel(' Erreur 3D (m) ')
plt.grid(True)
                              
plt.show()


          

