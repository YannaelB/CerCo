import numpy as np
import matplotlib as mpl
from openpyxl import load_workbook
from openpyxl import Workbook
import matplotlib.pyplot as plt
import os
from openpyxl.styles import Color,PatternFill
from numpy.linalg import solve
from collections import Counter
import statistics as st
from scipy import stats
import matplotlib.ticker as mticker


def ouvrirXLSX(filepath) :
    wb = load_workbook(filepath)
    return wb,wb.sheetnames

def lireFeuille(wb,nom_feuille,prem_ligne = 1,der_ligne=np.inf):
    ws = wb[nom_feuille]
    tableau = []
    der_ligne = min(der_ligne,ws.max_row)
    for i,row in enumerate(ws.rows,1) : # pour chaque ligne,
        if prem_ligne <= i <= der_ligne :
            ligne = []
            for cell in row : # pour chaque cellule de la ligne,
                ligne.append(cell.value) # lecture de son contenu
            tableau.append(ligne)
    tableau = np.array(tableau) # Conversion en tableau numpy
    shp,typ_don = tableau.shape,tableau.dtype
    return tableau



classeur = "Sujet_N_MutilAnalysis-nettoyé2.xlsx"


listeFeuilles = [lireFeuille(ouvrirXLSX(classeur)[0],i) for i in (ouvrirXLSX(classeur)[1])]




def etude_stimuli(L,n):
    TR = []
    Errh = []
    Errv = []
    Err3D = []
    for j in range(len(L)):
        for k in range(1,136):
            if int(L[j][k,3]) == n:
                if np.isnan(float(L[j][k,8])) == False:
                    TR.append(float(float(L[j][k,7])))
                    Errh.append(abs(float(L[j][k,9])))
                    Err3D.append(abs(float(L[j][k,8])))
                    Errv.append(abs(float(L[j][k,10])))

    return TR,Errh,Errv,Err3D
                    


    

   
Err_3D = []
for i in range(1,22):
    TR,Errh,Errv,Err3D = etude_stimuli(listeFeuilles,i)
    Err_3D.append(Err3D)

print("taille Err_3D : ",len(Err_3D))




### BoxPlot par sujet ###


#positions = [1,2,3,4,5,6,7,8,9,10,11,12,13,14],

plt.figure(4)
plt.violinplot([Err_3D[0],Err_3D[1],Err_3D[2],Err_3D[3],Err_3D[4],Err_3D[5],Err_3D[6],Err_3D[7],Err_3D[8],Err_3D[9],Err_3D[10],Err_3D[11],Err_3D[12],Err_3D[13],Err_3D[14],Err_3D[15],Err_3D[16],Err_3D[17],Err_3D[18],Err_3D[19],Err_3D[20]],points=300, widths=0.8, 
                     showmeans=True, showextrema=True, showmedians=True)

label_format = '{:,.0f}'
positions = [1, 2, 3, 4, 5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21]
labels = ['S1', 'S2','S3','S4','S5','S6','S7','S8', 'S9','S10','S11','S12','S13','S14','S15','S16','S17','S18','S19','S20','S21']
#plt.gca().xaxis.set_ticklabels(['S1', 'S2','S3','S4','S5','S6','S7','S8', 'S9','S10','S11','S12','S13','S14'])
#label_format = '{:.1%}'
plt.gca().xaxis.set_major_locator(mticker.FixedLocator(positions))
plt.gca().xaxis.set_major_formatter(mticker.FixedFormatter(labels))
plt.grid(True)
plt.title('Erreur 3D selon les stimuli')
plt.ylabel(' Erreur 3D (m) ')


                              
plt.show()


          

