import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
import matplotlib.pyplot as plt
import os
from openpyxl.styles import Color,PatternFill
from numpy.linalg import solve
from collections import Counter
import statistics as st
from scipy import stats


def ouvrirXLSX(filepath) :
    wb = load_workbook(filepath)
    return wb,wb.sheetnames

def lireFeuille(wb,nom_feuille,prem_ligne = 1,der_ligne=np.inf):
    ws = wb[nom_feuille]
    tableau = []
    der_ligne = min(der_ligne,ws.max_row)
    for i,row in enumerate(ws.rows,1) : # pour chaque ligne,
        if prem_ligne <= i <= der_ligne :
            ligne = []
            for cell in row : # pour chaque cellule de la ligne,
                ligne.append(cell.value) # lecture de son contenu
            tableau.append(ligne)
    tableau = np.array(tableau) # Conversion en tableau numpy
    shp,typ_don = tableau.shape,tableau.dtype
    return tableau



classeur = "Sujet_N_MutilAnalysis-nettoyé3.1.xlsx"  #chemin du fichier nettoyé par python


listeFeuilles = [lireFeuille(ouvrirXLSX(classeur)[0],i) for i in (ouvrirXLSX(classeur)[1])] #converti l'excel en une liste de tableau 




def etude_stimuli(L,R):
    TR = []
    Errh = []
    Errv = []
    Errp = []
    Err3D = []
    ErrHead = []
    for j in range(len(L)):
        #print('sujet numéro : ',j)
        for k in range(1,136):
            if int(L[j][k,2]) == R:
                if np.isnan(float(L[j][k,8])) == False:
                    TR.append(float(float(L[j][k,7])))  #créé des listes contenant les TR/Erreur azimuth/Erreur verticale/erreur profondeur/ etc.
                    Errh.append(abs(float(L[j][k,9])))
                    Errv.append(abs(float(L[j][k,10])))
                    Errp.append(abs(float(L[j][k,11])))
                    Err3D.append(abs(float(L[j][k,8])))
                if np.isnan(float(L[j][k,12])) == False:
                    ErrHead.append(abs(float(L[j][k,12])))
                   

    return TR,Errh,Err3D #pour alléger le programme, la fonction ne retourne que 3 listes. Pour ploter la donnée que vous voulez, il suffit de modifier "Err3D" par ce que vous voulez
#exemple : return TR,Errh,Err3D devient => return TR,Errh,Errv : cela plotera des graphiques de l'erreur verticale. Il suffit alors de changer les titres du graphique plot et le tour est joué !!
                    


    
                        
#on créé des listes pour la scène plage/church et grise

TR,Errh,Err3D = etude_stimuli(listeFeuilles,8) 
print("la taille de Err_3D est : ",len(Err3D)) #ces print ne servent qu'à vérifier/connaitre la taille des listes (très utile pour débugger)

TR2,Errh2,Err3D2 = etude_stimuli(listeFeuilles,10)
print("la taille de Err_3D est : ",len(Err3D2))

TR3,Errh3,Err3D3 = etude_stimuli(listeFeuilles,0)
print("la taille de Err_3D est : ",len(Err3D3))

    


        
### Trace un graphique BoxPlot des Err3D dans la plate/eglise/grise ###

plt.figure(1)
boxplotElements = plt.boxplot([Err3D,Err3D2,Err3D3], sym = 'b.', whis = 1.5,
                                 widths = [0.9,0.9,0.9], positions = [1,2,3],
                                 patch_artist = True)
plt.gca().xaxis.set_ticklabels(['plage', 'église','grise'])
for element in boxplotElements['medians']:
    element.set_color('blue')
    element.set_linewidth(1)
for element in boxplotElements['boxes']:
    element.set_edgecolor('magenta')
    element.set_facecolor('yellow')
    element.set_linewidth(1)
    element.set_linestyle('dashed')
    element.set_fill(True)
    element.set_hatch('/')
for element in boxplotElements['whiskers']:
    element.set_color('red')
    element.set_linewidth(1)
for element in boxplotElements['caps']:
    element.set_color('red')
plt.title('erreur selon les scènes (tout son confondu)') #titre du graph
plt.ylabel(' erreur profondeur (m) ')  #titre ordonnée graph
plt.grid(True)  #afficher les grilles/repère

plt.show()



