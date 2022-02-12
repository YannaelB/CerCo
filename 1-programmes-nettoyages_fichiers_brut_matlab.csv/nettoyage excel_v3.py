import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
import matplotlib.pyplot as plt
import os
from openpyxl.styles import Color,PatternFill
from numpy.linalg import solve
from collections import Counter
import math

def ouvrirXLSX(filepath) :  #cette fonction permet d'ouvrir un fichier excel selon un chemin donné
    wb = load_workbook(filepath)
    return wb,wb.sheetnames

#Cette fonction permet de parcourir le fichier excel et de le copier/coller dans un tableau pour le manipuler plus simplement. 
def lireFeuille(wb,nom_feuille,prem_ligne = 1,der_ligne=np.inf):   
    ws = wb[nom_feuille]
    tableau = []
    der_ligne = min(der_ligne,ws.max_row)
    for i,row in enumerate(ws.rows,1) : # pour chaque ligne,
        if prem_ligne <= i <= der_ligne :
            ligne = []
            for cell in row : # pour chaque cellule de la ligne,
                ligne.append(cell.value) # lecture de son contenu
            tableau.append(ligne)
    tableau = np.array(tableau) # Conversion en tableau numpy
    shp,typ_don = tableau.shape,tableau.dtype
    return tableau




classeur_data = 'Sujet_N_MutilAnalysis.xlsx'  #fichier données brut sorties de matlab (une feuille par sujet)
classeur_stim = 'stim_sujet_N.xlsx'  #fichier regroupant les documents stim (une feuille par sujet)


# creation d'une liste de tous les tableaux contenants les différents sujets/essais
listeFeuilles_data = [lireFeuille(ouvrirXLSX(classeur_data)[0],i) for i in (ouvrirXLSX(classeur_data)[1])]
listeFeuilles_stim = [lireFeuille(ouvrirXLSX(classeur_stim)[0],i) for i in (ouvrirXLSX(classeur_stim)[1])]



#Fonction qui supprime les 5 premières essais qui sont des essais d'entrainement pour les participants.
def suppr_5lignes(L):
    for i in range(len(L)):
        k = 0
        while int(L[i][1,1]) <= 5:
            L[i] = np.delete(L[i], 1, axis=0)
            k = k+1


#Fonction qui permet de calculer l'erreur entre l'angle de rotation de la tete et la position de l'enceinte (en azimuth)
            ## ATTENTION, je ne suis pas sûr de cette fonction, par manque de données, j'ai du supposer que Y est l'axe verticale, Z l'axe vers lequel le participant regarde et X l'axe passant par les 2 oreilles. DOnc l'azimuth est le plan (O,Z,X)
            ## Cependant, si l'on regarde bien les données de positions X du speaker, elles ne sont quasiment jamais négatives et varient très peu alors que je m'assurer bien que l'enceinte se trouve régulièrement sur toutes les positions prédéfinies
            ## ATTENTION, l'on ne peut pas se contenter de convertir les positions (A,B,C,etc.) en degrée car je ne respectais pas toujours l'ordre des positions indiquées par l'ordinateur afin de gagner du temps et  de pouvoir plus facilement jouer sur l'elevation
def head_angle(L):
    for i in range(len(L)):
        angle_speaker_2 = ['Speaker_azimuth_2']
        erreur_tete_2 = ['head_error_h']
        x = math.nan
        for j in range(1,len(L[i][:,1])): #permet de parcourir toutes les lignes d'une feuille en commencant par la 2nd
            if np.isnan(float(L[i][j,32])) == False: #permet de vérifier que la valeur existe (parfois il y a des bugs de Vive et des essais sont NaN)
                if float(L[i][j,34]) >= 0:
                    teta_speaker_2 = math.degrees(np.arctan(float(L[i][j,32])/float(L[i][j,34]))) #calcul la position azimuth du speaker grâce aux coordonnées cartésiennes => sphériques
                    angle_speaker_2.append(teta_speaker_2)
                    erreur_tete_2.append(abs(teta_speaker_2-float(L[i][j,30]))) #erreur égale absolu de la différence d'angle azimuth
                elif float(L[i][j,34])< 0:  #comme la tête ne peut pas pivoter de 180° à l'arriere, l'on ne prend pas en compte toutes les positions où le speaker se trouve derrière le participant
                    angle_speaker_2.append(x)
                    erreur_tete_2.append(x)
            else:
                angle_speaker_2.append(x)
                erreur_tete_2.append(x)
            
        L[i] = np.insert(L[i],52,erreur_tete_2,axis=1) #permet d'insérer la colonne d'erreur d'orientation de la tête dans l'excel brut à la 53ème colonne
        
            
            
                
#fonction qui permet de supprimer les colonnes inutiles du fichier brut          
def suppr_column(L):
    Suppr = list([2,4,5,6,7,8,9,10,11,12,13,14,15,17,19,20,21,22,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37,39,40,41,42,43,44,45,46,47,48])
    for i in range(len(L)):
        L[i] = np.delete(L[i], Suppr, axis=1)        

#fonction qui permet de compter les occurences et de supprimer toutes les répétitions afin d'alléger les fichiers bruts
def nettoyage_excel(L,M):
    for i in range(len(L)):
        column_Trial_i = L[i][1:,1]
        column_StimSound_i = M[i][:,8]
        column_Room_i = M[i][:,2]
        nb_move = []
        for j in range(6,141):
            nb_move.append(Counter(column_Trial_i)[str(j)]) # compte le nombre mvt tête (les occurences) et ajoute cela dans une liste

        m=1     
        while m <=135:
            if nb_move[m-1] != 1:
                for n in range(nb_move[m-1]-1):
                    L[i] = np.delete(L[i], m+1, axis=0)  #supprime les occurences inutiles
                m = m+1
            else:
                m = m + 1
                        

        column_StimSound_i = np.delete(column_StimSound_i,([1,2,3,4,5,6,7,53,99]),0) #supprimer les cases inutiles des fichiers stim
        column_Room_i = np.delete(column_Room_i,([1,2,3,4,5,6,7,53,99]),0)

        nb_move.insert(0,'nb_head_move')
        
        L[i] = np.insert(L[i],2,column_Room_i,axis=1)
        L[i] = np.insert(L[i],3,column_StimSound_i,axis=1)
        L[i] = np.insert(L[i],5,nb_move,axis=1)

    #toutes ces lignes permettent de supprimer les valeurs "absurdes" ou les essais pour lesquelles il y a eu des problèmes lors des manips
    L[0][12,5:],L[0][90,5:],L[0][129,5:] = math.nan,math.nan,math.nan
    L[2][57,5:],L[2][81,5:],L[2][120,5:] = math.nan,math.nan,math.nan
    L[3][12,5:],L[3][55,5:],L[3][62,5:] = math.nan,math.nan,math.nan
    L[4][13,5:] = math.nan
    L[5][24,5:],L[5][63,5:],L[5][82,5:],L[5][98,5:] = math.nan,math.nan,math.nan,math.nan
    L[6][39,5:],L[6][101,5:] = math.nan,math.nan
    L[7][26,5:],L[7][101,5:],L[7][116,5:],L[7][124,5:],L[7][130,5:] = math.nan,math.nan,math.nan,math.nan,math.nan  #supprimer le trial 31+106+121+129+135
    L[8][27,5:],L[8][42,5:],L[8][93,5:],L[8][96,5:],L[8][130,5:] = math.nan,math.nan,math.nan,math.nan,math.nan #sujet 9 (donc 8 python) : suppr trial 32+69+42?+98+101+135?+
    L[9][65,5:],L[9][82,5:],L[9][83,5:],L[9][120,5:],L[9][122,5:] = math.nan,math.nan,math.nan,math.nan,math.nan #sujet 10 (donc 9) : suppr trial 21?+70+87?+88+105?+107?+125+127+
    L[10][53,5:],L[10][58,5:],L[10][67,5:],L[10][105,5:] = math.nan,math.nan,math.nan,math.nan #sujet 11 (donc 10): suppr trial 16?+32?+58+63+72+110+132?+
    L[11][18,5:],L[11][75,5:],L[11][102,5:] = math.nan,math.nan,math.nan  #sujet 12 (donc 11): suppr trial 38?+49?+80+105?+107+
    L[12][75,5:],L[12][135,5:] = math.nan,math.nan  #sujet 13 (donc 12): suppr trial 22?+59?+65?+73?+80?+99?+140
    L[13][4,5:],L[13][14,5:],L[13][46,5:],L[13][57,5:],L[13][95,5:],L[13][101,5:],L[13][122,5:],L[13][125,5:],L[13][130,5:] = math.nan,math.nan,math.nan,math.nan,math.nan,math.nan,math.nan,math.nan,math.nan #sujet 14 (donc 13): suppr trial 9?+19+62+100+106+127+130+135?+


#Cette fonction permet de parcourir tout le fichier final afin de voir s'il n'y a pas des valeurs "absurdes" comme des erreurs 3D supérieures à 1m. Ensuite, il faut
#supprimer les essais à la main car il faut avant vérifier les essais détecter. Voir si cela correspond à un bug matlab, un bug de bases stations, un prroblème lors de la manip ou juste une grosse erreur du participant en question
def Détection_trials_suspects(L):
    for j in range(len(L)):
        for k in range(1,136):
            if abs(float(L[j][k,8])) >= 1:
                print("ATTENTION  !! Tu as une erreur_3D supérieure à 1m pour le sujet numéro: ", j+1," à l'essaie numéro  : ", k+5)
            if abs(float(L[j][k,9])) >= 100:
                print("ATTENTION  !! Tu as une erreur_asimuth supérieure à 100° pour le sujet numéro :", j+1," à l'essaie numéro : ", k+5)
     


#Cette fonction permet de transformer la liste de tableau en un fichier excel avec des feuilles
def creationwb(L):
    wb=Workbook()
    for i in range(len(L)):
        ws1 = wb.create_sheet("sujet_{}".format(i+1))
        
        # on rempli le tableau cellule par cellule
        for ligne in range (1,L[i].shape[0]+1) :
            for colonne in range (1,L[i].shape[1]+1):
                ws1.cell(row= ligne, column = colonne ,value = L[i][ligne-1][colonne-1] )
                
    del wb["Sheet"]
    wb.save("Sujet_N_MutilAnalysis-nettoyé3.xlsx")
    return print("l'excel à été créé ;) ")
        


suppr_5lignes(listeFeuilles_data)
head_angle(listeFeuilles_data)
suppr_column(listeFeuilles_data)
nettoyage_excel(listeFeuilles_data,listeFeuilles_stim)
Détection_trials_suspects(listeFeuilles_data)
creationwb(listeFeuilles_data)







